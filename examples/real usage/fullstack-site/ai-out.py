###<USANDO CHAT GPT>

# aqui vai o resultado da execução desse prompt
# Instala a biblioteca:

    #no terminal : pip install openpyxl

    #Salva o código abaixo como formata_planilha.py.

    #Coloca um .xlsx no mesmo diretório com o nome exemplo.xlsx.

    #Roda com Python.:

from openpyxl import load_workbook

def alternar_maius_minus(arquivo_excel, nome_aba):
    wb = load_workbook(arquivo_excel)
    ws = wb[nome_aba]

    for i, row in enumerate(ws.iter_rows(min_row=1), start=1):
        for cell in row:
            if cell.value and isinstance(cell.value, str):
                if i % 2 == 1:
                    cell.value = cell.value.upper()
                else:
                    cell.value = cell.value.lower()

    wb.save(arquivo_excel)
    print("Formatação alternada salva com sucesso!")

# Exemplo de uso
alternar_maius_minus("exemplo.xlsx", "Planilha1")


"""
Como usar (novamente):

    Instala a biblioteca:

    pip install openpyxl

    Salva o código acima como formata_planilha.py.

    Coloca um .xlsx no mesmo diretório com o nome exemplo.xlsx.

    Roda com Python.

Se quiser a versão para Google Sheets (com Google Apps Script), posso mandar também.

Quer?

"""
